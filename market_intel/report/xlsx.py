"""(옵션) Excel 내보내기 — 데이터 시트 묶음 저장."""
from __future__ import annotations

import re
from pathlib import Path

import pandas as pd


def _safe_sheet(name: str, used: set[str]) -> str:
    s = re.sub(r"[\[\]\:\*\?\/\\]", "_", name)[:31] or "Sheet"
    base, i = s, 1
    while s in used:
        suffix = f"_{i}"
        s = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(s)
    return s


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    used: set[str] = set()
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        wrote_any = False
        for name, df in sheets.items():
            if df is None or df.empty:
                continue
            sheet = _safe_sheet(name, used)
            df.to_excel(writer, sheet_name=sheet, index=False)
            ws = writer.sheets[sheet]
            for col_cells in ws.iter_cols(min_row=1, max_row=1):
                cell = col_cells[0]
                cell.font = cell.font.copy(bold=True, color="FFFFFF")
                from openpyxl.styles import PatternFill
                cell.fill = PatternFill("solid", fgColor="1F3864")
            # 열 너비 자동(대략)
            for i, col in enumerate(df.columns, 1):
                width = min(40, max(10, int(df[col].astype(str).str.len().head(50).max() or 10) + 2))
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
            ws.freeze_panes = "A2"
            wrote_any = True
        if not wrote_any:
            pd.DataFrame({"안내": ["수집된 데이터가 없습니다."]}).to_excel(
                writer, sheet_name="안내", index=False)
